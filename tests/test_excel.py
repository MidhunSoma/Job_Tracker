import os
import tempfile
from pathlib import Path
from sqlalchemy.orm import Session
import openpyxl
from job_tracker.config.settings import settings
from job_tracker.models.application import Application, ApplicationStatus
from job_tracker.services.excel_service import ExcelService


def test_excel_sync(db_session: Session):
    """Verifies that the Excel sync service writes applications and styles columns correctly."""
    fd, temp_path_str = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    temp_path = Path(temp_path_str)
    
    original_path = settings.export.excel_path
    settings.export.excel_path = temp_path
    
    try:
        app1 = Application(
            company="Google",
            role="AI Engineer",
            status=ApplicationStatus.TECHNICAL_INTERVIEW,
            recruiter_name="Jane Doe",
            recruiter_email="jane@google.com",
            gmail_link="https://mail.google.com/1",
            notes="Completed coding challenge successfully"
        )
        app2 = Application(
            company="Microsoft",
            role="Software Engineer",
            status=ApplicationStatus.APPLIED,
            recruiter_name="John Smith",
            recruiter_email="john@microsoft.com",
            gmail_link="https://mail.google.com/2",
            notes="Applied through referral"
        )
        db_session.add_all([app1, app2])
        db_session.commit()
        
        service = ExcelService()
        service.sync_db_to_excel(db_session)
        
        assert temp_path.exists()
        wb = openpyxl.load_workbook(temp_path)
        ws = wb.active
        
        assert ws.title == "Job Applications"
        assert ws.max_row == 3
        
        headers = [ws.cell(row=1, column=col).value for col in range(1, 10)]
        assert "Company" in headers
        assert "Role" in headers
        assert "Current Status" in headers
        
        row2_company = ws.cell(row=2, column=3).value
        row3_company = ws.cell(row=3, column=3).value
        companies = {row2_company, row3_company}
        assert "Google" in companies
        assert "Microsoft" in companies
        assert ws.views.sheetView[0].showGridLines is True
        
    finally:
        settings.export.excel_path = original_path
        if temp_path.exists():
            os.remove(temp_path)
