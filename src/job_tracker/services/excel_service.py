import openpyxl
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from ..config.settings import settings
from ..repositories.application import ApplicationRepository
from ..utils.logging import logger


class ExcelService:
    """Service handling synchronization of SQLite database records to the Excel tracker workbook."""

    def __init__(self) -> None:
        pass

    @property
    def filepath(self) -> Path:
        path = settings.export.excel_path
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    def sync_db_to_excel(self, db: Session) -> None:
        """Reads all job applications from SQLite and overwrites/synchronizes them to the Excel workbook.

        Applies professional formatting: bold headers, frozen first row, auto-fitted columns,

        gridlines, and enables data filters.
        """
        logger.info(
            f"Syncing SQLite applications database with Excel at {self.filepath}...",
            extra={"action": "EXCEL_SYNC_START"}
        )
        
        try:
            # 1. Fetch all applications from DB
            app_repo = ApplicationRepository(db)
            applications = app_repo.get_all()
            
            # Sort applications: latest updated first (or by applied date)
            # Sorting by last_updated descending for relevance
            applications = sorted(applications, key=lambda a: a.last_updated, reverse=True)

            # 2. Create a new workbook or load existing
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Job Applications"
            
            # Ensure gridlines are visible
            ws.views.sheetView[0].showGridLines = True

            # 3. Define headers
            headers = [
                "Date Applied",
                "Last Updated",
                "Company",
                "Role",
                "Current Status",
                "Recruiter",
                "Recruiter Email",
                "Mail Link",
                "Notes"
            ]
            ws.append(headers)

            # 4. Populate rows
            for app in applications:
                applied_str = app.applied_date.strftime("%Y-%m-%d") if app.applied_date else ""
                updated_str = app.last_updated.strftime("%Y-%m-%d %H:%M:%S") if app.last_updated else ""
                
                row_data = [
                    applied_str,
                    updated_str,
                    app.company,
                    app.role,
                    app.status,
                    app.recruiter_name or "",
                    app.recruiter_email or "",
                    app.gmail_link or "",
                    app.notes or ""
                ]
                ws.append(row_data)

            # 5. Apply styling and formatting
            # Styles definitions
            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Sleek Dark Indigo
            align_center = Alignment(horizontal="center", vertical="center", wrap_text=False)
            align_left = Alignment(horizontal="left", vertical="center", wrap_text=False)
            align_wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            thin_border = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9")
            )
            
            # Style header row (Row 1)
            ws.row_dimensions[1].height = 28
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border

            # Style data rows
            font_data = Font(name="Segoe UI", size=11)
            for row_idx in range(2, ws.max_row + 1):
                ws.row_dimensions[row_idx].height = 20
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.font = font_data
                    cell.border = thin_border
                    
                    # Alignments
                    if col_idx in [1, 2]:  # Dates
                        cell.alignment = align_center
                    elif col_idx == 9:  # Notes (wrap text for readability)
                        cell.alignment = align_wrap_left
                    else:
                        cell.alignment = align_left

            # 6. Freeze first row
            ws.freeze_panes = "A2"

            # 7. Enable Filters
            if ws.max_row > 1:
                ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

            # 8. Auto-size column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                
                # Check lengths of values in this column
                for cell in col:
                    # Ignore Notes cell length for width fitting (prevent overly wide Notes column)
                    if cell.column == 9:
                        continue
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                
                # Set dynamic width
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                
            # Explicitly set Notes column width
            ws.column_dimensions[get_column_letter(9)].width = 40

            # 9. Save workbook
            wb.save(self.filepath)
            logger.info(
                f"Excel sync successful. Wrote {len(applications)} records.",
                extra={"action": "EXCEL_SYNC_SUCCESS", "records_count": len(applications)}
            )
            
        except Exception:
            logger.exception("Failed to write SQLite records to Excel.")
            raise
